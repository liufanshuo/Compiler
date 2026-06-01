from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_OUTPUT_SUFFIX = "_性能得分"
FINAL_SCORE_HEADER = "最终性能测试成绩"


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def find_default_input() -> Path:
    candidates = [
        path
        for path in Path.cwd().glob("*.xlsx")
        if not path.name.startswith("~$")
        and DEFAULT_OUTPUT_SUFFIX not in path.stem
        and "性能得分" not in path.stem
    ]
    if len(candidates) != 1:
        names = ", ".join(path.name for path in candidates) or "none"
        raise FileNotFoundError(
            "Please specify an input file with --input. "
            f"Auto-detected candidates: {names}"
        )
    return candidates[0]


def copy_cell_style(source, target) -> None:
    if not source.has_style:
        return
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def detect_columns(ws) -> tuple[int, int, int, int]:
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    try:
        rank_col = headers.index("rank") + 1
    except ValueError as exc:
        raise ValueError("Cannot find required column: rank") from exc

    try:
        name_col = headers.index("姓名") + 1
    except ValueError:
        name_col = 3

    test_start_col = name_col + 1
    test_end_col = rank_col - 1
    final_score_col = rank_col + 1

    if test_start_col > test_end_col:
        raise ValueError("Cannot detect benchmark/test-point columns.")
    if final_score_col > ws.max_column:
        raise ValueError("Cannot find the final score output column after rank.")

    return test_start_col, test_end_col, rank_col, final_score_col


def competition_ranks(scores: dict[int, float | None]) -> dict[int, int | None]:
    valid = [(row, score) for row, score in scores.items() if score is not None]
    valid.sort(key=lambda item: item[1], reverse=True)

    ranks: dict[int, int | None] = {row: None for row in scores}
    previous_score: float | None = None
    previous_rank = 0
    for index, (row, score) in enumerate(valid, start=1):
        if previous_score is None or score != previous_score:
            previous_rank = index
            previous_score = score
        ranks[row] = previous_rank
    return ranks


def calculate_scores(input_path: Path, output_path: Path, sheet_name: str | None = None) -> None:
    wb = load_workbook(input_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    test_start_col, test_end_col, rank_col, final_score_col = detect_columns(ws)
    data_start_row = 2
    data_end_row = ws.max_row

    final_scores: dict[int, float | None] = {}

    for col in range(test_start_col, test_end_col + 1):
        positive_times: list[float] = []
        for row in range(data_start_row, data_end_row + 1):
            value = ws.cell(row=row, column=col).value
            if is_number(value) and value > 0:
                positive_times.append(float(value))

        min_time = min(positive_times) if positive_times else None

        for row in range(data_start_row, data_end_row + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if min_time is None or not is_number(value) or value <= 0:
                cell.value = None
            else:
                cell.value = round(100.0 * min_time / float(value), 6)
                cell.number_format = "0.000000"

    for row in range(data_start_row, data_end_row + 1):
        row_scores = [
            ws.cell(row=row, column=col).value
            for col in range(test_start_col, test_end_col + 1)
            if is_number(ws.cell(row=row, column=col).value)
        ]
        final_scores[row] = (
            round(sum(float(score) for score in row_scores) / len(row_scores), 6)
            if row_scores
            else None
        )

    ranks = competition_ranks(final_scores)

    header_source = ws.cell(row=1, column=final_score_col)
    ws.cell(row=1, column=final_score_col).value = FINAL_SCORE_HEADER
    copy_cell_style(header_source, ws.cell(row=1, column=final_score_col))

    for row in range(data_start_row, data_end_row + 1):
        rank_cell = ws.cell(row=row, column=rank_col)
        score_cell = ws.cell(row=row, column=final_score_col)
        rank_cell.value = ranks[row]
        score_cell.value = final_scores[row]
        score_cell.number_format = "0.000000"

    for col in range(test_start_col, test_end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(
            ws.column_dimensions[get_column_letter(col)].width or 10,
            12,
        )
    ws.column_dimensions[get_column_letter(final_score_col)].width = max(
        ws.column_dimensions[get_column_letter(final_score_col)].width or 10,
        18,
    )

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Calculate benchmark performance scores from a contest ranking workbook."
    )
    parser.add_argument("-i", "--input", type=Path, help="Input .xlsx file. Defaults to the only source .xlsx in current directory.")
    parser.add_argument("-o", "--output", type=Path, help="Output .xlsx file. Defaults to '<input>_性能得分.xlsx'.")
    parser.add_argument("-s", "--sheet", help="Worksheet name. Defaults to the active sheet.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = args.input or find_default_input()
    output_path = args.output or input_path.with_name(f"{input_path.stem}{DEFAULT_OUTPUT_SUFFIX}.xlsx")

    calculate_scores(input_path, output_path, args.sheet)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
